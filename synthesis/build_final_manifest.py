# -*- coding: utf-8 -*-
"""One machine receipt for the state the bundle is actually in.

WHY A NEW ONE. Three receipts described three different moments and none
described this one: the integrated-status manifest was written against the
1,724-effect table before the figure rows, the repaired round-6 JSON sees
1,728 rows but validates only the ten sheet/CSV pairs it knew about, and the
completion note is prose. Meanwhile two of the twelve pairs had gone stale
without anything noticing, because the check that reported "all matched" had
only ten pairs in its own definition of "all". A receipt whose scope is fixed
by hand will keep saying everything matched while the thing it does not look
at drifts.

So the pair list here is DERIVED, not typed: every sheet in the workbook that
has a CSV counterpart on disk is compared, and the count of pairs is an output
of the run rather than an assumption in it. A thirteenth CSV-backed sheet would
appear here on its own.

FAIL-CLOSED. This exits non-zero if any pair disagrees, if a formula has lost
its cached result, if the source-corpus receipt no longer accounts for every
artifact on disk, or if any human-only field has been filled in. It is meant to
be the thing that refuses, not the thing that reassures.
"""
import csv, hashlib, io, json, os, subprocess, sys, collections
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import bundle_paths
BASE = bundle_paths.BASE
WB = os.path.join(BASE, 'extraction_form_RASi_PCa.xlsx')
FAILURES = []


def fail(msg):
    FAILURES.append(msg)


def sha(p):
    h = hashlib.sha256()
    with open(p, 'rb') as f:
        for b in iter(lambda: f.read(65536), b''):
            h.update(b)
    return h.hexdigest()


def read_rows(path):
    d = '\t' if path.endswith('.tsv') else ','
    with io.open(path, encoding='utf-8-sig', newline='') as fh:
        return [r for r in csv.reader(fh, delimiter=d) if any(c != '' for c in r)]


R = {'generated': '2026-08-31', 'bundle': os.path.basename(BASE)}

# ---------------------------------------------------------------- workbook
wbf = openpyxl.load_workbook(WB, data_only=False)
wbv = openpyxl.load_workbook(WB, data_only=True)
formulas = [(w.title, c.coordinate) for w in wbf.worksheets
            for row in w.iter_rows() for c in row
            if isinstance(c.value, str) and c.value.startswith('=')]
lost = [s + '!' + c for s, c in formulas if wbv[s][c].value is None]
cached_err = [s + '!' + c for s, c in formulas
              if isinstance(wbv[s][c].value, str) and wbv[s][c].value.startswith('#')]
R['workbook'] = {'sheets': len(wbf.sheetnames), 'formulas': len(formulas),
                 'formulas_without_cached_value': lost,
                 'cached_error_cells': cached_err, 'sha256': sha(WB)}
if lost:
    fail('%d formulas have lost their cached value' % len(lost))
if cached_err:
    fail('%d cached error cells' % len(cached_err))

# ------------------------------------------------- the pairs, DERIVED not typed
CANDIDATES = {}
for name in sorted(os.listdir(BASE)):
    if name.endswith(('.csv', '.tsv')):
        CANDIDATES[name] = os.path.join(BASE, name)
KNOWN = {'effects_text_long': 'effect_extraction_text_long.csv',
         'extraction_counts_long': 'extraction_counts_long.csv',
         'descriptive_statistics_long': 'extraction_descriptive_long.csv',
         'text_remaining_work': 'text_extraction_remaining_work.csv',
         'csv_data_dictionary': 'extraction_data_dictionary.csv',
         'text_extraction_status': 'text_extraction_status.csv',
         'extraction': 'extraction_form_RASi_PCa.csv',
         'fulltext_screening': 'fulltext_screening.tsv',
         'source_retrieval_queue': 'text_source_retrieval_queue.csv',
         'text_source_coverage': 'text_table_coverage.csv',
         'operational_status_all_eligible': 'extraction_operational_status.csv',
         'source_corpus_receipt': 'source_corpus_receipt.csv'}
pairs, unpaired = {}, []
for sheet in wbf.sheetnames:
    csvname = KNOWN.get(sheet)
    if not csvname:
        # a sheet whose name matches a CSV stem is a pair nobody declared
        for n in CANDIDATES:
            if os.path.splitext(n)[0] == sheet:
                csvname = n
                break
    if not csvname:
        continue
    if csvname not in CANDIDATES:
        fail('sheet %s expects %s, which is not on disk' % (sheet, csvname))
        continue
    d = read_rows(CANDIDATES[csvname])
    rows = [r for r in wbv[sheet].iter_rows(values_only=True)
            if any(c not in (None, '') for c in r)]
    bad = 0
    if len(rows) == len(d):
        for i in range(len(d)):
            xr = [('' if c is None else str(c)) for c in rows[i][:len(d[0])]]
            for a, b in zip(xr, d[i]):
                if a != b:
                    try:
                        if float(a) == float(b):
                            continue
                    except Exception:
                        pass
                    bad += 1
    pairs[sheet] = {'csv': csvname, 'xlsx_rows': len(rows), 'csv_rows': len(d),
                    'rowcount_match': len(rows) == len(d), 'cell_diffs': bad,
                    'csv_sha256': sha(CANDIDATES[csvname])}
    if len(rows) != len(d) or bad:
        fail('pair %s <-> %s disagrees (%d rows vs %d, %d cell diffs)'
             % (sheet, csvname, len(rows), len(d), bad))
R['sheet_csv_pairs'] = pairs
R['pair_count'] = len(pairs)
R['sheets_without_csv'] = [s for s in wbf.sheetnames if s not in pairs]

# ------------------------------------------------------------- long tables
def dictrows(name):
    d = '\t' if name.endswith('.tsv') else ','
    with io.open(os.path.join(BASE, name), encoding='utf-8-sig', newline='') as fh:
        return list(csv.DictReader(fh, delimiter=d))


eff = dictrows('effect_extraction_text_long.csv')
cnt = dictrows('extraction_counts_long.csv')
desc = dictrows('extraction_descriptive_long.csv')
R['long_tables'] = {
    'effect_rows': len(eff), 'count_rows': len(cnt), 'descriptive_rows': len(desc),
    'records_with_effects': len({r['rec_id'] for r in eff}),
    'derivation_method': dict(collections.Counter(r['derivation_method'] for r in eff)),
    'rows_added_2026_08_30': sum(1 for r in eff if r['ai_correction_date'] == '2026-08-30')}

# -------------------------------------------------------------- human gates
gates = {'effect_human_confirmed': dict(collections.Counter(r['human_confirmed'] for r in eff)),
         'effect_pool_eligible': dict(collections.Counter(r['pool_eligible'] for r in eff)),
         'count_human_confirmed': dict(collections.Counter(r['human_confirmed'] for r in cnt)),
         'screening_human_confirmed': dict(collections.Counter(
             r['human_confirmed'] for r in dictrows('fulltext_screening.tsv')))}
ws = wbv['extraction']
hdr = [c.value for c in ws[1]]
for field in ('extractor1_initials', 'extractor2_initials', 'rob_overall',
              'human_confirmed'):
    j = hdr.index(field) + 1
    vals = [ws.cell(row=i, column=j).value for i in range(2, ws.max_row + 1)]
    gates['extraction_' + field] = dict(collections.Counter(
        '' if v is None else str(v) for v in vals))
R['human_gates'] = gates
if gates['effect_pool_eligible'] != {'no': len(eff)}:
    fail('pool_eligible is not "no" on every effect row')
if set(gates['effect_human_confirmed']) - {'no'}:
    fail('human_confirmed is not "no" on every effect row')
for f in ('extractor1_initials', 'extractor2_initials', 'rob_overall'):
    if set(gates['extraction_' + f]) - {''}:
        fail('%s has been filled in; that field is a person\'s' % f)

# ------------------------------------------------------- the route gate itself
# THE FUNCTION CI VERIFIES IS THE FUNCTION THIS RECEIPT RUNS. The bundle's
# qc_extraction carries the same rule, and two copies of a rule drift: the
# public CI could stay green while the gate the bundle actually uses changed.
# So the receipt calls the public one directly, and if the two ever disagree
# this fails on the difference rather than on nobody noticing.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import route_gate

_routes = route_gate.route_index(dictrows('fulltext_screening.tsv'))
_route_findings = route_gate.findings(eff, _routes)
R['route_gate'] = {
    'records_with_a_route': len(_routes),
    'ambiguous_screening_routes': sorted(
        rid for rid, v in _routes.items()
        if v.get('_status') == route_gate.AMBIGUOUS),
    'findings': [{'effect_row_id': a, 'code': b, 'detail': c}
                 for a, b, c in _route_findings[:20]],
    'finding_count': len(_route_findings)}
if _route_findings:
    fail('%d effect rows disagree with the route screening gave their record'
         % len(_route_findings))

# ------------------------------------------------------------ source corpus
rec = dictrows('source_corpus_receipt.csv')
listed = {r['source_path'] for r in rec if r['source_path']}
disk = set()
for d in ('fulltext_pdfs', 'source_supplements'):
    p = os.path.join(BASE, d)
    if os.path.isdir(p):
        for f in sorted(os.listdir(p)):
            # README.txt is documentation; the other .txt files here are real
            # source artifacts - extracted full text for records that have no
            # PDF - and excluding every .txt dropped four of them.
            if not f.startswith('.') and f.lower() != 'readme.txt':
                disk.add('%s/%s' % (d, f))
R['source_corpus'] = {
    'rows': len(rec),
    'by_status': dict(collections.Counter(r['receipt_status'] for r in rec)),
    'records_covered': len({r['rec_id'] for r in rec if r['rec_id']}),
    'artifacts_on_disk': len(disk), 'artifacts_listed': len(listed),
    'on_disk_not_listed': sorted(disk - listed),
    'listed_not_on_disk': sorted(listed - disk)}
if disk - listed or listed - disk:
    fail('source corpus receipt does not account for every artifact on disk')

# ------------------------------------------------------------- figure rows
R['figure_transcribed_rows'] = [
    {k: r[k] for k in ('effect_row_id', 'rec_id', 'synthesis_readiness',
                       'analysis_stream', 'effect_measure', 'effect_point',
                       'effect_ci_low', 'effect_ci_high', 'ci_level',
                       'source_page_ref', 'discrepancy_flag', 'pool_eligible',
                       'human_confirmed')}
    for r in eff if r['data_source'] == 'figure (printed numbers)']
if len(R['figure_transcribed_rows']) != 4:
    fail('expected 4 figure-transcribed rows, found %d'
         % len(R['figure_transcribed_rows']))

# --------------------------------------------------------- refusals, guards
sup = os.path.join(BASE, 'outputs', 'supplement_integration_2026-08-30', 'logs')
for name, key in (('integration.json', 'refused_values'),):
    p = os.path.join(sup, name)
    if os.path.exists(p):
        R['refused_values'] = len(json.load(io.open(p, encoding='utf-8'))[key])
p = os.path.join(sup, 'R1087_coexposure_pending.json')
if os.path.exists(p):
    R['held_without_comparator'] = len(json.load(io.open(p, encoding='utf-8'))['printed_values'])
R['idempotency'] = {}
for n in ('integrate_supplements', 'figure_printed_numbers'):
    p = os.path.join(sup, '%s_idempotency.json' % n)
    if os.path.exists(p):
        R['idempotency'][n] = json.load(io.open(p, encoding='utf-8'))['verdict']
if set(R['idempotency'].values()) - {'ALREADY_PRESENT_NO_WRITE'}:
    fail('a writer would still append on a rerun: %s' % R['idempotency'])

# ------------------------------------------------------------------- QC
out = subprocess.run([sys.executable, 'qc_extraction.py',
                      'extraction_form_RASi_PCa.xlsx', '--out', 'qc_report.md'],
                     cwd=BASE, capture_output=True, text=True)
line = (out.stdout or '').strip().splitlines()[-1] if out.stdout else ''
qc = dict(kv.split('=') for kv in line.split() if '=' in kv)
tally = {}
for l in io.open(os.path.join(BASE, 'qc_report.md'), encoding='utf-8'):
    import re as _re
    m = _re.match(r'^\| (BLOCKER|WARN) \| ([^|]+)\| (\d+) \|', l)
    if m:
        tally['%s/%s' % (m.group(1), m.group(2).strip())] = int(m.group(3))
R['qc'] = {'summary': qc, 'by_check': tally,
           'exit_code_meaning': 'qc_extraction exits 1 while any BLOCKER stands'}
if qc.get('synthesis_ready') != '0':
    fail('synthesis_ready is not 0')

# ---------------------------------------------------------------- hashes
R['file_sha256'] = {n: sha(os.path.join(BASE, n))
                    for n in sorted(os.listdir(BASE))
                    if n.endswith(('.csv', '.tsv', '.xlsx', '.yaml'))}

R['verdict'] = 'CONSISTENT' if not FAILURES else 'FAILED'
R['failures'] = FAILURES
json.dump(R, io.open(os.path.join(HERE, 'FINAL_MANIFEST.json'), 'w',
                     encoding='utf-8'), indent=2, ensure_ascii=False)

print('시트 %d개 중 CSV 대응 %d쌍' % (R['workbook']['sheets'], R['pair_count']))
for s, d in pairs.items():
    print('  %-34s %5d/%-5d diffs=%d' % (s, d['xlsx_rows'], d['csv_rows'],
                                         d['cell_diffs']))
print('수식 %d / 캐시 손실 %d' % (R['workbook']['formulas'], len(lost)))
print('출처 영수증 %s' % R['source_corpus']['by_status'])
print('QC %s' % qc)
print('재실행 가드 %s' % R['idempotency'])
print()
print('판정: %s' % R['verdict'])
for f in FAILURES:
    print('   - %s' % f)
raise SystemExit(1 if FAILURES else 0)
