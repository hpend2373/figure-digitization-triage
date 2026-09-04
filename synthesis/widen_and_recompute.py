# -*- coding: utf-8 -*-
"""Widen the formula ranges over the effects table and refresh their cached
results - and nothing else.

Adding rows to a sheet does not extend the formulas that count it: a COUNTIF
written over $B$2:$B$1725 keeps stopping at 1725 and quietly ignores every row
after it. This widens such ranges to the sheet's real last row and recomputes
the cached result, so that a data_only reader and Excel agree.

TWO THINGS IT WILL NOT TOUCH.

  text_extraction_summary!B26 is labelled 'pre-supplement projection snapshot'
  and its 1609 bound is deliberate - a frozen before-picture, not a stale
  range. Widening it would destroy the only record of that state. Excluded by
  cell, explicitly.

  Any formula this cannot fully parse. An earlier version of this fix summed
  the COUNTIFs it recognised in a formula and returned that sum, so a formula
  with no COUNTIF at all - a direct reference, a ROUND - received the empty
  sum, zero, and 160 live cached values were overwritten with 0 before the
  report printed. Here an unparsed formula returns None and the cell is left
  exactly as found.
"""
import csv, io, os, re, shutil, sys, zipfile
import xml.sax.saxutils as SU
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import bundle_paths
BASE = bundle_paths.BASE
WB = os.path.join(BASE, 'extraction_form_RASi_PCa.xlsx')
FROZEN = {('text_extraction_summary', 'B26')}


def col_index(name):
    n = 0
    for ch in name:
        n = n * 26 + (ord(ch) - 64)
    return n - 1


with io.open(os.path.join(BASE, 'effect_extraction_text_long.csv'),
             encoding='utf-8-sig', newline='') as fh:
    rd = csv.DictReader(fh)
    cols, rows = rd.fieldnames, list(rd)
LAST = len(rows) + 1
print('effects_text_long 실제 마지막 행: %d' % LAST)


def countifs(terms):
    n = 0
    for r in rows:
        for letter, crit in terms:
            v = r[cols[col_index(letter)]]
            if crit == '<>':
                ok = v != ''
            elif crit.startswith('<>'):
                ok = v != crit[2:]
            else:
                ok = v == crit
            if not ok:
                break
        else:
            n += 1
    return n


RNG = r"'effects_text_long'!\$([A-Z]+)\$\d+:\$[A-Z]+\$(\d+)"
PAIR = re.compile(RNG + r'\s*,\s*(?:"([^"]*)"|\$?([A-Z]+)\$?(\d+))')
ONLY = re.compile(r'^=?COUNTA\(' + RNG + r'\)$')


def calls(f):
    out = []
    for m in re.finditer(r'COUNTIFS?\(', f):
        i, d = m.end(), 1
        while i < len(f) and d:
            d += (f[i] == '(') - (f[i] == ')')
            i += 1
        out.append((m.start(), i))
    return out


def recompute(f, resolve):
    m = ONLY.match(f.strip())
    if m:
        j = col_index(m.group(1))
        return sum(1 for r in rows if r[cols[j]] != '')
    sp = calls(f)
    if not sp:
        return None
    rest = f
    for a, b in reversed(sp):
        rest = rest[:a] + rest[b:]
    if re.sub(r'[\s+=]', '', rest):
        return None
    total = 0
    for a, b in sp:
        inner = f[a:b][f[a:b].index('(') + 1:-1]
        terms, end = [], 0
        for mm in PAIR.finditer(inner):
            lit, rc, rr = mm.group(3), mm.group(4), mm.group(5)
            crit = lit if lit is not None else resolve(rc, int(rr))
            if crit is None:
                return None
            terms.append((mm.group(1), str(crit)))
            end = mm.end()
        if not terms or inner[end:].strip():
            return None
        total += countifs(terms)
    return total


vals = openpyxl.load_workbook(WB, data_only=True)
with zipfile.ZipFile(WB) as z:
    wbx = z.read('xl/workbook.xml').decode('utf-8')
    rels = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')
    members = [(i, z.read(i.filename)) for i in z.infolist()]
raw = dict((i.filename, d) for i, d in members)


def part_of(name):
    m = re.search(r'<x:sheet name="%s"[^>]*r:id="([^"]+)"' % re.escape(name), wbx)
    t = (re.search(r'Target="([^"]+)"[^>]*Id="%s"' % m.group(1), rels)
         or re.search(r'Id="%s"[^>]*Target="([^"]+)"' % m.group(1), rels))
    return t.group(1).lstrip('/')


shutil.copy2(WB, os.path.join(BASE, 'archive',
                              'pre_supplement_integration_2026-08-30',
                              'extraction_form_RASi_PCa.pre_widen.xlsx'))
WIDEN = re.compile(r"('effects_text_long'!\$[A-Z]+\$2:\$[A-Z]+\$)(\d+)")
CELL = re.compile(r'<x:c r="([A-Z]+)(\d+)"([^>]*)><x:f>(.*?)</x:f><x:v>([^<]*)</x:v></x:c>')
edited, report, left = {}, [], []
for sheet in [n for n in vals.sheetnames]:
    sp = part_of(sheet)
    sx = raw[sp].decode('utf-8')
    if 'effects_text_long' not in sx:
        continue
    ws = vals[sheet]
    st = [0, 0, 0]

    def fix(m):
        c, rn, attrs, f, cached = m.groups()
        if 'effects_text_long' not in f:
            return m.group(0)
        if (sheet, c + rn) in FROZEN:
            st[2] += 1
            return m.group(0)
        f2 = WIDEN.sub(lambda s: s.group(1) + str(LAST), f)
        got = recompute(SU.unescape(f2), lambda cc, rr: ws['%s%d' % (cc, rr)].value)
        if got is None:
            left.append((sheet, c + rn))
            return m.group(0) if f2 == f else (
                '<x:c r="%s%s"%s><x:f>%s</x:f><x:v>%s</x:v></x:c>'
                % (c, rn, attrs, f2, cached))
        st[0] += 1
        if str(got) != cached:
            st[1] += 1
            report.append((sheet, c + rn, cached, str(got)))
        return ('<x:c r="%s%s"%s><x:f>%s</x:f><x:v>%s</x:v></x:c>'
                % (c, rn, attrs, f2, got))

    sx2 = CELL.sub(fix, sx)
    if sx2 != sx:
        edited[sp] = sx2.encode('utf-8')
    print('  %-26s 재계산 %3d, 값변경 %2d, 동결 %d' % (sheet, st[0], st[1], st[2]))

tmp = WB + '.tmp'
with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as out:
    for info, data in members:
        out.writestr(info, edited.get(info.filename, data))
shutil.move(tmp, WB)
print('\n손대지 않은 수식 %d개 (직접참조 등)' % len(left))
print('바뀐 캐시값 %d개:' % len(report))
for s, r, a, b in report:
    print('   %-26s %-5s %s -> %s' % (s, r, a, b))
